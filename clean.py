"""
clean.py - Providers Exclusion List: cleaning / ETL step
=========================================================
Reads the two raw source files:
    1. UPDATED.csv                          (Federal OIG LEIE)
    2. DCH_OIG_Exclusions_List_*.xlsx       (Georgia DCH OIG)

Applies the agreed cleaning rules:
    - ALL columns NOT NULL, missing data uses sentinel values:
        strings -> ''          dates -> 1900-01-01
    - placeholder identifiers ('0000000000') are dropped, real
      NPI/UPIN values become rows in identifiers.csv
    - dates parsed from YYYYMMDD, invalid/placeholder -> sentinel
    - STATE etc. trimmed + uppercased
    - party_type derived (INDIVIDUAL / ENTITY)
    - parties deduplicated per source; ids pre-assigned so that
      seed.py can bulk-insert with create_many (which does not
      return generated ids)
    - GA highlighted rows -> recently_added = True

Outputs (./out/):
    sources.csv, import_logs.csv, parties.csv,
    identifiers.csv, exclusions.csv
"""

import os
import re
from datetime import datetime, date

import pandas as pd
from openpyxl import load_workbook

# ---------------------------------------------------------------
# Config - adjust paths to your local files
# ---------------------------------------------------------------
FEDERAL_CSV = "raw_data/UPDATED.csv"
GA_XLSX = "raw_data/DCH_OIG_Exclusions_List_06092026.xlsx"
OUT_DIR = "out"

SENTINEL_DATE = "1900-01-01"
FEDERAL_SNAPSHOT = "2026-06-01"   # LEIE monthly snapshot date (adjust if known)
GA_SNAPSHOT = "2026-06-09"        # from the file name / title banner
IMPORT_DATE = date.today().isoformat()

SOURCE_FEDERAL = 1
SOURCE_GA = 2
IMPORT_FEDERAL = 1
IMPORT_GA = 2


# ---------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------
def clean_str(v) -> str:
    """Trim; None/NaN -> '' (string sentinel)."""
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() == "nan":
        return ""
    return s


def clean_date(v) -> str:
    """YYYYMMDD string -> ISO date; placeholder/invalid -> sentinel."""
    s = clean_str(v)
    s = re.sub(r"\.0$", "", s)          # excel numeric cells: '19820415.0'
    if not s or set(s) == {"0"}:
        return SENTINEL_DATE
    try:
        return datetime.strptime(s, "%Y%m%d").date().isoformat()
    except ValueError:
        print(f"  [warn] invalid date {s!r} -> sentinel")
        return SENTINEL_DATE


def clean_identifier(v) -> str:
    """Real identifier or '' (all-zero placeholders are dropped)."""
    s = clean_str(v)
    if not s or set(s) == {"0"}:
        return ""
    return s


def upper(v) -> str:
    return clean_str(v).upper()


# ---------------------------------------------------------------
# Accumulators (ids pre-assigned, seed.py inserts verbatim)
# ---------------------------------------------------------------
parties = []        # dicts with party_id
identifiers = []    # dicts with identifier_id
exclusions = []     # dicts with exclusion_id
party_index = {}    # dedup key -> party_id

def get_party_id(key, row_dict):
    """Return existing party_id for key, or register a new party."""
    if key in party_index:
        return party_index[key]
    pid = len(parties) + 1
    row_dict["party_id"] = pid
    parties.append(row_dict)
    party_index[key] = pid
    return pid


_seen_identifiers = set()

def add_identifier(party_id, id_type, value):
    if not value:
        return
    # avoid duplicate (party, type, value) rows - matches DB unique constraint
    k = (party_id, id_type, value)
    if k in _seen_identifiers:
        return
    _seen_identifiers.add(k)
    identifiers.append({
        "identifier_id": len(identifiers) + 1,
        "party_id": party_id,
        "identifier_type": id_type,
        "identifier_value": value,
    })


# ---------------------------------------------------------------
# 1. Federal LEIE (UPDATED.csv)
# ---------------------------------------------------------------
print("Reading federal LEIE ...")
df = pd.read_csv(FEDERAL_CSV, dtype=str, keep_default_na=False)
print(f"  {len(df)} rows")

for r in df.itertuples(index=False):
    last = upper(r.LASTNAME)
    first = upper(r.FIRSTNAME)
    mid = upper(r.MIDNAME)
    bus = upper(r.BUSNAME)
    dob = clean_date(r.DOB)
    addr = upper(r.ADDRESS)

    if last:
        ptype = "INDIVIDUAL"
        key = ("F-I", last, first, mid, dob)
    else:
        ptype = "ENTITY"
        key = ("F-E", bus, addr)

    pid = get_party_id(key, {
        "party_type": ptype,
        "last_name": last,
        "first_name": first,
        "middle_name": mid,
        "business_name": bus,
        "dob": dob,
        "address": addr,
        "city": upper(r.CITY),
        "state": upper(r.STATE),
        "zip_code": clean_str(r.ZIP),
    })

    add_identifier(pid, "NPI", clean_identifier(r.NPI))
    add_identifier(pid, "UPIN", clean_identifier(r.UPIN))

    exclusions.append({
        "exclusion_id": len(exclusions) + 1,
        "party_id": pid,
        "source_id": SOURCE_FEDERAL,
        "import_id": IMPORT_FEDERAL,
        "general_category": upper(r.GENERAL),
        "specialty": upper(r.SPECIALTY),
        "exclusion_type": clean_str(r.EXCLTYPE),
        "exclusion_date": clean_date(r.EXCLDATE),
        "reinstatement_date": clean_date(r.REINDATE),
        "waiver_date": clean_date(r.WAIVERDATE),
        "waiver_state": upper(r.WVRSTATE),
        "status": "ACTIVE",
        "recently_added": False,
    })

federal_count = len(exclusions)
print(f"  -> {federal_count} exclusion records, {len(parties)} parties so far")

# ---------------------------------------------------------------
# 2. Georgia DCH list (xlsx) - read with openpyxl so that data
#    and highlight (recently added) info stay row-aligned.
#    Header row is Excel row 3; data starts at row 4.
# ---------------------------------------------------------------
print("Reading Georgia DCH list ...")
wb = load_workbook(GA_XLSX)
ws = wb["Sheet1"]

NO_FILL = (None, "00000000", "FFFFFFFF")
ga_rows = 0

for excel_row in range(4, ws.max_row + 1):
    cells = [ws.cell(row=excel_row, column=c).value for c in range(1, 9)]
    # LAST, FIRST, MIDDLE, BUSINESS, GENERAL, STATE, SANCDATE, NPI
    if all(clean_str(c) == "" for c in cells):
        continue  # skip blank rows
    ga_rows += 1

    last, first, mid = upper(cells[0]), upper(cells[1]), upper(cells[2])
    bus = upper(cells[3])

    fill = ws.cell(row=excel_row, column=1).fill
    rgb = fill.fgColor.rgb if fill and fill.fgColor else None
    highlighted = rgb not in NO_FILL

    # 7 GA rows have both a person name and a business name;
    # classify as ENTITY whenever a business name is present.
    if bus:
        ptype = "ENTITY"
        key = ("G-E", bus)
    else:
        ptype = "INDIVIDUAL"
        key = ("G-I", last, first, mid)

    pid = get_party_id(key, {
        "party_type": ptype,
        "last_name": last,
        "first_name": first,
        "middle_name": mid,
        "business_name": bus,
        "dob": SENTINEL_DATE,   # GA list has no DOB
        "address": "",          # GA list has no address fields
        "city": "",
        "state": upper(cells[5]),
        "zip_code": "",
    })

    npi = clean_identifier(cells[7])
    if npi:
        npi = re.sub(r"\.0$", "", npi)
        if not re.fullmatch(r"\d{10}", npi):
            print(f"  [warn] row {excel_row}: non-standard NPI {npi!r} skipped")
            npi = ""
    add_identifier(pid, "NPI", npi)

    exclusions.append({
        "exclusion_id": len(exclusions) + 1,
        "party_id": pid,
        "source_id": SOURCE_GA,
        "import_id": IMPORT_GA,
        "general_category": upper(cells[4]),
        "specialty": "",            # not provided by GA list
        "exclusion_type": "",       # not provided by GA list
        "exclusion_date": clean_date(cells[6]),   # SANCDATE
        "reinstatement_date": SENTINEL_DATE,
        "waiver_date": SENTINEL_DATE,
        "waiver_state": "",
        "status": "ACTIVE",
        "recently_added": highlighted,
    })

ga_count = len(exclusions) - federal_count
print(f"  -> {ga_count} exclusion records "
      f"({sum(e['recently_added'] for e in exclusions)} recently added)")

# ---------------------------------------------------------------
# 3. Source / import-log metadata
# ---------------------------------------------------------------
sources = [
    {"source_id": SOURCE_FEDERAL, "source_name": "OIG LEIE",
     "source_level": "FEDERAL", "file_name": os.path.basename(FEDERAL_CSV),
     "file_type": "CSV", "source_date": FEDERAL_SNAPSHOT},
    {"source_id": SOURCE_GA, "source_name": "GA DCH OIG",
     "source_level": "STATE", "file_name": os.path.basename(GA_XLSX),
     "file_type": "XLSX", "source_date": GA_SNAPSHOT},
]
import_logs = [
    {"import_id": IMPORT_FEDERAL, "source_id": SOURCE_FEDERAL,
     "import_date": IMPORT_DATE, "records_imported": federal_count,
     "notes": "Initial import of federal LEIE snapshot"},
    {"import_id": IMPORT_GA, "source_id": SOURCE_GA,
     "import_date": IMPORT_DATE, "records_imported": ga_count,
     "notes": "Initial import of GA DCH snapshot; highlighted rows "
              "flagged as recently_added"},
]

# ---------------------------------------------------------------
# 4. Write cleaned CSVs
# ---------------------------------------------------------------
os.makedirs(OUT_DIR, exist_ok=True)
pd.DataFrame(sources).to_csv(f"{OUT_DIR}/sources.csv", index=False)
pd.DataFrame(import_logs).to_csv(f"{OUT_DIR}/import_logs.csv", index=False)
pd.DataFrame(parties).to_csv(f"{OUT_DIR}/parties.csv", index=False)
pd.DataFrame(identifiers).to_csv(f"{OUT_DIR}/identifiers.csv", index=False)
pd.DataFrame(exclusions).to_csv(f"{OUT_DIR}/exclusions.csv", index=False)

print("\nSummary")
print(f"  parties:     {len(parties)}")
print(f"  identifiers: {len(identifiers)}")
print(f"  exclusions:  {len(exclusions)}")
print(f"Cleaned files written to ./{OUT_DIR}/")
